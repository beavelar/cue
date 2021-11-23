import logging
from pandas.core.frame import DataFrame
from openpyxl.chart.marker import Marker
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import Series, Reference, ScatterChart

#########################################################################################################
# Logger definition

logging.basicConfig(
    format="%(levelname)s: %(asctime)s - %(name)s.%(funcName)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

#########################################################################################################


def generate_excel_body(worksheet: Worksheet, data: DataFrame) -> None:
    """
    generate_excel_body
    ----------

    This function will set the body of a excel sheet to the elements from the dataframe
    """
    logger.info("Generating the body for the excel spread sheet")
    for index in range(data.Ticker.size):
        cell_number = str(index + 2)
        worksheet[f"A{cell_number}"] = data.Ticker[index]
        worksheet[f"B{cell_number}"] = data["Option Type"][index]
        worksheet[f"C{cell_number}"] = data["Alerted At"][index]
        worksheet[f"D{cell_number}"] = data["Day of Week"][index]
        worksheet[f"E{cell_number}"] = data["Time of Day"][index]
        worksheet[f"F{cell_number}"] = data.Expiry[index]
        worksheet[f"G{cell_number}"] = data["Days to Exp."][index]
        worksheet[f"H{cell_number}"] = data.Strike[index]
        worksheet[f"I{cell_number}"] = data.Underlying[index]
        worksheet[f"J{cell_number}"] = data["Diff %"][index]
        worksheet[f"K{cell_number}"] = data.Volume[index]
        worksheet[f"L{cell_number}"] = data["Open Interest"][index]
        worksheet[f"M{cell_number}"] = data["Vol/OI"][index]
        worksheet[f"N{cell_number}"] = data["Implied Volatility"][index]
        worksheet[f"O{cell_number}"] = data.Delta[index]
        worksheet[f"P{cell_number}"] = data.Gamma[index]
        worksheet[f"Q{cell_number}"] = data.Vega[index]
        worksheet[f"R{cell_number}"] = data.Theta[index]
        worksheet[f"S{cell_number}"] = data.Rho[index]
        worksheet[f"T{cell_number}"] = data["Alert Ask"][index]
        worksheet[f"U{cell_number}"] = data["Highest Ask"][index]
        worksheet[f"V{cell_number}"] = data["P/L"][index]
        worksheet[f"W{cell_number}"] = data["Time Passed"][index]


#########################################################################################################


def create_excel_header(worksheet: Worksheet) -> None:
    """
    create_excel_header
    ----------

    This function will set the header of a excel sheet
    """
    logger.info("Creating the header for the excel spread sheet")
    style_excel_header(worksheet)
    worksheet["A1"] = "Ticker"
    worksheet["B1"] = "Option Type"
    worksheet["C1"] = "Alerted At"
    worksheet["D1"] = "Day of Week"
    worksheet["E1"] = "Time of Day"
    worksheet["F1"] = "Expiry"
    worksheet["G1"] = "Days to Exp."
    worksheet["H1"] = "Strike"
    worksheet["I1"] = "Underlying"
    worksheet["J1"] = "Diff %"
    worksheet["K1"] = "Volume"
    worksheet["L1"] = "Open Interest"
    worksheet["M1"] = "Vol/OI"
    worksheet["N1"] = "Implied Volatility"
    worksheet["O1"] = "Delta"
    worksheet["P1"] = "Gamma"
    worksheet["Q1"] = "Vega"
    worksheet["R1"] = "Theta"
    worksheet["S1"] = "Rho"
    worksheet["T1"] = "Alert Ask"
    worksheet["U1"] = "Highest Ask"
    worksheet["V1"] = "P/L"
    worksheet["W1"] = "Time Passed"


#########################################################################################################


def style_excel_header(worksheet: Worksheet) -> None:
    """
    style_excel_header
    ----------

    This function will set the style of the header of a excel sheet
    """
    logger.info("Styling the header for the excel spread sheet")
    black_fill = PatternFill(
        start_color="00000000", end_color="00000000", fill_type="solid"
    )

    font = Font(color="FFFFFFFF")

    cells = worksheet["A1":"W1"][0]
    for cell in cells:
        cell.font = font
        cell.fill = black_fill
        cell.alignment = Alignment(horizontal="center")


#########################################################################################################


def set_worksheet_settings(worksheet: Worksheet) -> None:
    """
    set_worksheet_settings
    ----------

    This function will set the settings of a excel sheet
    """
    logger.info("Setting the worksheet settings")
    worksheet.title = "Historical Alerts"
    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 21
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 12
    worksheet.column_dimensions["F"].width = 12
    worksheet.column_dimensions["G"].width = 12
    worksheet.column_dimensions["I"].width = 11
    worksheet.column_dimensions["J"].width = 9
    worksheet.column_dimensions["L"].width = 13
    worksheet.column_dimensions["M"].width = 12
    worksheet.column_dimensions["N"].width = 16
    worksheet.column_dimensions["U"].width = 11
    worksheet.column_dimensions["V"].width = 10
    worksheet.column_dimensions["W"].width = 12


#########################################################################################################


def create_scatter_chart(x_title: str, y_title: str) -> ScatterChart:
    """
    create_scatter_chart
    ----------

    This function will create a scatter chart with the provided x-axis and y-axis titles
    """
    logger.info("Creating scatter chart from the provided data")
    chart = ScatterChart()
    chart.title = f"{x_title} to {y_title}"
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.scatterStyle = "marker"
    chart.style = 13
    return chart


#########################################################################################################


def generate_chart_series(
    worksheet: Worksheet,
    chart: ScatterChart,
    x_column: int,
    y_column: int,
    num_of_elems: int,
) -> None:
    """
    generate_chart_series
    ----------

    This function will generate the series from the columns provided and append it to the chart
    """
    logger.info("Generating the chart series for the desired chart")
    x_axis = Reference(worksheet, min_col=x_column, min_row=2, max_row=num_of_elems)
    y_axis = Reference(worksheet, min_col=y_column, min_row=1, max_row=num_of_elems)
    series = Series(y_axis, x_axis, title_from_data=True)
    chart.series.append(series)


#########################################################################################################


def style_chart(chart: ScatterChart) -> None:
    """
    style_chart
    ----------

    This function will set the styling properties of the provided chart
    """
    logger.info("Setting the style of the provided chart")
    style = chart.series[0]
    style.marker = Marker("circle")
    style.marker.size = 3
    style.graphicalProperties.line.noFill = True
    style.marker.graphicalProperties.solidFill = "000000"  # Marker filling
    style.marker.graphicalProperties.line.solidFill = "000000"  # Marker outline


#########################################################################################################


def create_excel_chart(
    file: str,
    x_title: str,
    y_title: str,
    x_column: int,
    y_column: int,
    graph_cell: str,
    num_of_elems: int,
) -> None:
    """
    create_excel_chart
    ----------

    This function will create a scatter chart and set it in the provided cell
    """
    logger.info("Creating the desired excel chart")
    file_wb = load_workbook(file)
    worksheet = file_wb.active
    chart = create_scatter_chart(x_title, y_title)
    generate_chart_series(worksheet, chart, x_column, y_column, num_of_elems)
    style_chart(chart)
    worksheet.add_chart(chart, graph_cell)
    file_wb.save(file)


#########################################################################################################


def save_df_to_excel(dataframe: DataFrame, path: str) -> None:
    """
    save_df_to_excel
    ----------

    This function will save the data provided by the dataframe onto the provided excel path
    """
    logger.info("Saving the dataframe to the desired excel file")
    output_wb = Workbook()
    worksheet = output_wb.active

    create_excel_header(worksheet)
    generate_excel_body(worksheet, dataframe)
    set_worksheet_settings(worksheet)
    output_wb.save(path)


#########################################################################################################


def color_cell(worksheet: Worksheet, cell: str, color: str) -> None:
    """
    color_cell
    ----------

    This function will color the desired cell to the desired color
    """
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    worksheet_cell = worksheet[cell]
    worksheet_cell.fill = fill
